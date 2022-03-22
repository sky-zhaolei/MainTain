from openpyxl import load_workbook
import os

#获取数据当前目录下的测试数据文件
file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "login_datas.xlsx")

#1、加载excel数据文件
wb = load_workbook(file_path)
#读取对应的表单
sh = wb["login"]
#获取表单中单元格的数据，单元格对象sh.cell(row,colum)   openpyxl中下标默认从1开始
#获取当前表单中的总行数（sh.max_row），总列数(max_colum)
#修改单元格数据 sh.cell(row,colum).value = 新的值 ,修改后记得保存整个文件（wb.save(文件路径)）
#保存到原文件时，需要注意文件不能被占用，否则会报错,sh.rows读取当前表中所有的数据
#list（sh.rows) 读取出来，每一行是个元组，元组放的时每一个行的单元格
#excel文件读取出来的数据：字符串，数字

data_lists = []

#获取首行字段名称值，即字典key值
titles = []        #读取首行
for item in list(sh.rows)[0]:
    titles.append(item.value)    #将获取到的首行，加入到titles列表

#获取除首行外的其他行所有数据
for item in list(sh.rows)[1:]:
    value_dict = []  #每一行是个新字典
    for index in item:
        value_dict.append(index.value)   #将读取到的每一行值添加入value_dict列表
    res = dict(zip(titles,value_dict))   #将titles和读取出的每行数据，均打包成字典
    res["check"] = eval(res["check"])    #在追加到列表中时，将check行转换为字符对象

    data_lists.append(res)  #将每一行字段均追加到data_lists列表中
print(data_lists)






